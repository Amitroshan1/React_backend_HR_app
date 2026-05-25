"""Helpers for expense claim attachment paths and Excel export."""

from io import BytesIO
from pathlib import Path

CLAIM_EXPENSES_SUBDIR = "expenses"
EXPENSE_CLAIM_TEMPLATE = (
    Path(__file__).resolve().parent / "templates" / "Expenses_Claim_Form_FT.xlsx"
)


def claim_attach_static_filename(stored_name):
    """
    Path segment for url_for('static', filename=...).
    Files on disk: website/static/uploads/expenses/<name>
    """
    if not stored_name:
        return None
    name = (stored_name or "").replace("\\", "/").strip().lstrip("/")
    if name.startswith(f"{CLAIM_EXPENSES_SUBDIR}/"):
        return f"uploads/{name}"
    return f"uploads/{CLAIM_EXPENSES_SUBDIR}/{name}"


def claim_attach_storage_name(basename):
    """Relative path stored in ExpenseLineItem.Attach_file (under uploads/expenses/)."""
    base = (basename or "").replace("\\", "/").strip().lstrip("/")
    if not base:
        return None
    if base.startswith(f"{CLAIM_EXPENSES_SUBDIR}/"):
        return base
    return f"{CLAIM_EXPENSES_SUBDIR}/{base}"


def generate_expense_claim_excel(header, line_items, *, circle=None, emp_type=None, claim_status=None):
    """
    Fill official Expenses_Claim_Form_FT.xlsx (same layout as Expenses_Claim_Form_Bengaluru sample).
    """
    from openpyxl import load_workbook

    if not EXPENSE_CLAIM_TEMPLATE.is_file():
        raise FileNotFoundError(f"Expense claim template not found: {EXPENSE_CLAIM_TEMPLATE}")

    def fmt_date_dash(d):
        if not d:
            return ""
        return d.strftime("%d-%m-%Y")

    def fmt_date_colon(d):
        if not d:
            return ""
        return d.strftime("%d:%m:%Y")

    def clear_advance_payment_block(worksheet):
        """Remove template sample advance amounts (e.g. 124000 in Q22)."""
        for r in range(16, 22):
            for col in (16, 17, 18):  # P Date, Q Amount, R Remark
                try:
                    worksheet.cell(row=r, column=col, value=None)
                except AttributeError:
                    pass
        worksheet["Q22"] = None
        for col in (16, 17, 18):
            try:
                worksheet.cell(row=22, column=col, value=None)
            except AttributeError:
                pass

    wb = load_workbook(EXPENSE_CLAIM_TEMPLATE)
    ws = wb.active
    clear_advance_payment_block(ws)

    header_date = header.travel_to_date or header.travel_from_date
    if not header_date and line_items:
        for li in line_items:
            if getattr(li, "date", None):
                header_date = li.date
                break

    ws["D7"] = (header.employee_name or "").strip()
    ws["K7"] = fmt_date_dash(header_date) if header_date else ""
    ws["D8"] = (header.designation or "").strip()
    ws["D9"] = (header.emp_id or "").strip()
    ws["D10"] = (header.project_name or "").strip()

    place = (header.country_state or "").strip()
    ws["C11"] = place or "—"
    if header.travel_from_date:
        ws["D11"] = f"Date:{fmt_date_colon(header.travel_from_date)}"
    ws["H11"] = "To:"
    ws["I11"] = place or "—"
    if header.travel_to_date:
        ws["J11"] = f"Date: {fmt_date_colon(header.travel_to_date)}"

    items = list(line_items or [])
    data_start = 14
    template_slots = 10
    n = len(items)
    extra = max(0, n - template_slots)
    if extra:
        ws.insert_rows(24, extra)
    sum_row = 24 + extra
    advance_row = sum_row + 2
    net_row = sum_row + 4

    for r in range(data_start, sum_row):
        for col in (2, 3, 4, 8, 9, 10, 11, 12, 13):
            ws.cell(row=r, column=col, value=None)

    for idx, li in enumerate(items):
        r = data_start + idx
        ws.cell(row=r, column=2, value=li.sr_no)
        ws.cell(row=r, column=3, value=fmt_date_dash(li.date) if li.date else "")
        ws.cell(row=r, column=4, value=(li.purpose or "").strip())
        ws.cell(
            row=r,
            column=8,
            value="Printed Receipts" if li.Attach_file else "No Receipts",
        )
        amt = float(li.amount or 0)
        cur = (li.currency or "INR").strip().upper()
        if cur == "USD":
            ws.cell(row=r, column=11, value=amt)
        elif cur in ("EUR", "EURO"):
            ws.cell(row=r, column=12, value=amt)
        else:
            ws.cell(row=r, column=13, value=amt)

    last_item_row = data_start + n - 1 if n else data_start
    total_inr = 0.0
    for li in items:
        cur = (li.currency or "INR").strip().upper()
        if cur in ("INR", ""):
            total_inr += float(li.amount or 0)

    if n:
        ws[f"J{sum_row}"] = f"=SUM(J{data_start}:J{last_item_row})"
        ws[f"K{sum_row}"] = f"=SUM(K{data_start}:K{last_item_row})"
        ws[f"L{sum_row}"] = f"=SUM(L{data_start}:L{last_item_row})"
        ws[f"M{sum_row}"] = f"=SUM(M{data_start}:M{last_item_row})"

    for col in (10, 11, 12, 13):  # J, K, L, M — Advance Taken (B) left blank
        ws.cell(row=advance_row, column=col, value=None)

    ws[f"J{net_row}"] = f"=J{sum_row}-IF(ISBLANK(J{advance_row}),0,J{advance_row})"
    ws[f"K{net_row}"] = f"=K{sum_row}-IF(ISBLANK(K{advance_row}),0,K{advance_row})"
    ws[f"L{net_row}"] = f"=L{sum_row}-IF(ISBLANK(L{advance_row}),0,L{advance_row})"
    if n:
        ws[f"M{net_row}"] = f"=M{sum_row}-IF(ISBLANK(M{advance_row}),0,M{advance_row})"
        if total_inr:
            ws[f"M{net_row}"] = round(total_inr, 2)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
